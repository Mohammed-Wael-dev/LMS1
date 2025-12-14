from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django_tenants.utils import tenant_context

from iraq_form.models import VotingCenter
from tenant.models import Client


class Command(BaseCommand):
    help = "Seed voting centers into a specific tenant schema."

    def add_arguments(self, parser):
        parser.add_argument(
            "schema_name",
            type=str,
            help="Tenant schema where the voting centers will be created.",
        )
        parser.add_argument(
            "--file",
            dest="file_path",
            default="مراكز الاقتراع.xlsx",
            help="Path to the Excel file that contains the voting center names.",
        )

    def handle(self, *args, **options):
        schema_name = options["schema_name"]
        file_path = options["file_path"]

        names = self._load_names(file_path)
        if not names:
            raise CommandError("No voting center names were found in the provided file.")

        try:
            tenant = Client.objects.get(schema_name=schema_name)
        except Client.DoesNotExist as exc:  # pragma: no cover
            raise CommandError(f"Tenant with schema '{schema_name}' does not exist.") from exc

        created_count = 0

        with tenant_context(tenant):
            for name in names:
                _, created = VotingCenter.objects.get_or_create(name=name)
                if created:
                    created_count += 1
                    self.stdout.write(f"Added voting center: {name}")
                else:
                    self.stdout.write(f"Voting center already exists: {name}")

        skipped_count = len(names) - created_count

        self.stdout.write(
            self.style.SUCCESS(
                f"Inserted {created_count} voting centers into schema '{schema_name}'."
            )
        )

        if skipped_count:
            self.stdout.write(
                self.style.WARNING(
                    f"Skipped {skipped_count} existing voting centers."
                )
            )

    def _load_names(self, file_path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise CommandError("openpyxl is required to read Excel files. Please install it.") from exc

        path = Path(file_path)
        if not path.is_absolute():
            path = Path(settings.BASE_DIR) / path

        if not path.exists():
            raise CommandError(f"File '{path}' was not found.")

        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover
            raise CommandError(f"Could not open Excel file '{path}'.") from exc

        sheet = workbook.active
        names = []

        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            cell_value = row[0]
            if cell_value is None:
                continue
            name = str(cell_value).strip()
            if not name:
                continue
            if index == 1 and name == "المدارس":
                continue
            names.append(name)

        workbook.close()

        names = list(dict.fromkeys(names))

        return names
